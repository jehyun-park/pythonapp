import openpyxl

wb = openpyxl.Workbook() # 엑셀 만들기

ws = wb.create_sheet('업체별 현재시세') # 워크시트 만들기

#데이터추가
ws['A1'] = '업체명'
ws['B1'] = '현재시세'

ws['A2'] = '카카오'
ws['A3'] = '네이버'
ws['A4'] = '삼성전자'
ws['A5'] = 'SK하이닉스'
ws['A6'] = 'LG전자'

wb.save('업체별 현재시세.xlsx')